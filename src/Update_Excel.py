from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from email_service import send_with_outlook
FIELDS = [
    "timestamp", "wake_word", "utterance", "recognized_text", "intent",
    "is_final_asr","cpu_usage" ,"prompt_text",
     "confidence","retry_count", "error_code"
]

EXCEL_FILE = f"session_log_{datetime.now().strftime('%Y-%m-%d_%H_%M_%S')}.xlsx"

class Update_Excel():
    _instance = None
    def __init__(self):
        self.data = {key: "" for key in FIELDS}
        self.data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#**kwargs = "Keyword Arguments" It allows the function to accept any number of named arguments, even ones that aren’t pre-defined.
    def update(self, **kwargs): 
        for key, value in kwargs.items():
            if key in self.data:
                self.data[key] = value

    def write(self, file_name=EXCEL_FILE):
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Log"
            ws.append(FIELDS)
        else:
            wb = load_workbook(file_name)
            ws = wb["Log"]

        row = [self.data.get(field, "") for field in FIELDS]
        ws.append(row)
        wb.save(file_name)
        print(f"✔ Session written to {file_name}")

    def reset(self):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.data = {key: "" for key in FIELDS}
        self.data["timestamp"] = timestamp
    def send_mail():
        print("sending mail through win32 client after tts")
        send_with_outlook(EXCEL_FILE)
if __name__ == "__main__":
    session = Update_Excel()
    #  Wake word detected
    session.update(wake_word="Hey Reva")
    #  ASR result comes in later
    session.update(asr_text="Turn on lights", intent="Control_Lights", confidence="0.91")
    # System info received
    session.update(cpu_usage="12%", mem_usage="155MB")
    # Response generated
    session.update(response_text="Turning on the lights", retry_count="0")
    # Finally write once everything is ready
    session.write()
    #  Start next session
    session.reset()
